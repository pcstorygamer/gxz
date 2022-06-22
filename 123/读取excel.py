# -*- coding:utf-8 -*-

from openpyxl import Workbook,load_workbook
from openpyxl.styles import *

import warnings
warnings.filterwarnings('ignore')

wb = load_workbook('台湾资源类型整理.xlsx')
# print(wb.sheetnames)
ws = wb['统计']
# print(ws.values)
# for i in ws.values:
#     print(i)
#

# for i in ws['A']:
#     # print(i.value,i)
#     if "9088" == str(i.value):
#         print(i)

sss = "A"+"888"
print(ws[sss].value)